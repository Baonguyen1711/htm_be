from fastapi import APIRouter, HTTPException, File, UploadFile
from fastapi import FastAPI, Depends, UploadFile
from starlette.requests import Request
from firebase_admin import auth
from ..database import db
from ..services.auth_service import verify_token
from google.cloud import firestore
from collections import defaultdict
from ..models.questions import UpdateQuestionRequest, Answer, Grid, PlacementArray
from ..models.scores import Score, ScoreRule
from fastapi.encoders import jsonable_encoder
from openpyxl import load_workbook
from io import BytesIO
from typing import Optional, List
import logging
import traceback
from app.stores.player_store import get_player_info
from app.services.firestore_service import upload_test_to_firestore, get_test_by_name, get_test_name_by_user_id, update_question, upload_single_question_to_firestore
from ..services.test_service import process_test_data, get_specific_question, get_questions_by_round, get_packet_name
from ..services.realtime_service import send_selected_packet_name_to_player,is_existing_player,get_all_player_answer,set_single_player_answer,send_question_to_player, send_answer_to_player, start_time, broadcast_player_answer, send_score, send_round_2_grid_to_player, send_selected_row_to_player, send_incorrect_row_to_player, send_correct_row_to_player, send_obstacle, send_packet_name_to_player, send_current_question_to_player, send_selected_cell_to_player, send_cell_color_to_player, send_score_rule, update_score_each_round, get_current_correct_answer, get_player_answer, get_player_answer_correct, set_current_correct_answer, set_player_answer, set_player_answer_correct, get_score_rules
import re


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

test_routers = APIRouter()

player_answer = defaultdict(list)
score_rules = defaultdict(dict)


def normalize_string(text: str) -> str:
    # Lowercase
    text = text.lower()
    # Remove leading/trailing spaces
    text = text.strip()
    # Replace multiple spaces/tabs/newlines with a single space
    text = re.sub(r'\s+', ' ', text)
    return text

@test_routers.put("/api/test/update/{question_id}")
def update_question_document(question_id, request: UpdateQuestionRequest):
    try:
        logger.info("Attempting to create reference to Firebase")
        update_data = jsonable_encoder(request)
        logger.info(f"updated_data {update_data}")
        logger.info("get question correctly")
        
    
    except Exception as e:
        logger.error(f"Error creating Firebase reference: {str(e)}")
        logger.error(f"Full traceback: {traceback.format_exc()}")
        raise 

    if not request:
        raise HTTPException(status_code=400, detail="No fields provided to update.")
    
    try:
        logger.info("Attempting to create reference to Firebase")
        result = update_question(question_id, update_data)
        logger.info(f"updated_data {update_data}")
        logger.info("get question correctly")
        
    
    except Exception as e:
        logger.error(f"Error creating Firebase reference: {str(e)}")
        logger.error(f"Full traceback: {traceback.format_exc()}")
        raise 
    # result = update_question(question_id, update_data)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result

@test_routers.get("/api/test/user")
def read_test_by_user_id(request: Request):
    user = request.state.user
    authenticated_uid = user["uid"]
    logger.info(f"authenticated_uid: {authenticated_uid}")

    # if user_id != authenticated_uid:
    #     raise HTTPException(status_code=403, detail="Unauthorized access to this user's tests")
    
    test_list = get_test_name_by_user_id(authenticated_uid)
    if "error" in test_list:
        raise HTTPException(status_code=500, detail=test_list["error"])
    if len(test_list) == 0:
        raise HTTPException(status_code=404, detail="no test found")
    return test_list

@test_routers.post("/api/test/grid/color")
def set_selected_cell(room_id: str,row_index:str, col_index:str,color:str,request: Request):

    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        try:
            logger.info("Attempting to create reference to Firebase")
            send_cell_color_to_player(room_id,row_index,col_index, color)
            logger.info("get question correctly")
            
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@test_routers.post("/api/test/grid/cell")
def set_cell_color(room_id: str,row_index:str, col_index:str,request: Request):

    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        try:
            logger.info("Attempting to create reference to Firebase")
            send_selected_cell_to_player(room_id,row_index,col_index)
            logger.info("get question correctly")
            
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@test_routers.post("/api/test/grid")
def send_grid_to_player(room_id: str,request: Request, grid: Grid):
    # player_answer[f"{room_id}"] = []
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        try:
            logger.info("Attempting to create reference to Firebase")
            send_round_2_grid_to_player(grid.grid, room_id)
            logger.info("get question correctly")
            
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/row/action")
def set_selected_column(room_id: str, row_number: str, action:str, request: Request,  word_length: int,correct_answer: Optional[str] = None ,marked_characters_index: Optional[str] = None, is_row:Optional[bool] = None):
    # player_answer[f"{room_id}"] = []
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        try:
            if action == "SELECT":
                logger.info("Attempting to create reference to Firebase")
                send_selected_row_to_player(room_id, row_number,is_row, word_length)
                logger.info("get question correctly")
            if action == "CORRECT":
                logger.info(f"marked_character_index {marked_characters_index}")
                send_correct_row_to_player(room_id, row_number, correct_answer, marked_characters_index, is_row, word_length)
                logger.info("get question correctly")
            if action == "INCORRECT":
                logger.info("Attempting to create reference to Firebase")
                send_incorrect_row_to_player(room_id, row_number, is_row,word_length)
                logger.info("get question correctly")
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/row/open")
def set_selected_column(room_id: str, row_number: str, request: Request):
    # player_answer[f"{room_id}"] = []
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        try:
            logger.info("Attempting to create reference to Firebase")
            send_selected_row_to_player(room_id, row_number)
            logger.info("get question correctly")
            
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/obstacle")
async def open_obstacle(room_id: str, obstacle: str ,request: Request):

    try:
        # user = request.state.user
        # authenticated_uid = user["uid"]
        try:
            body = await request.json()
            logger.info(f"Received body: {body}")
        
            placementArray = body  # since you expect a list
            logger.info("Attempting to create reference to Firebase")

            send_obstacle(room_id, obstacle,placementArray)
            logger.info("obstacle opened")
            
        
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@test_routers.post("/api/test/question/current")
def get_each_round_questions(room_id:str,question_number:int,request: Request):

    try:
        # user = request.state.user

        # authenticated_uid = user["uid"]
        send_current_question_to_player(room_id,question_number)

        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/packet/set")
def set_selected_packet(packet_name: str,room_id:str,request: Request):

    try:
        user = request.state.user
        authenticated_uid = user["uid"]


        try:
            logger.info("Attempting to create reference to Firebase")

            send_selected_packet_name_to_player(packet_name,room_id)
            logger.info("get question correctly")

            return packet_name
            
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@test_routers.get("/api/test/question/round/packet")
def get_each_round_questions(test_name:str ,room_id:str,request: Request):

    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        result = process_test_data(authenticated_uid, test_name)

        try:
            logger.info("Attempting to create reference to Firebase")
            packets = get_packet_name(result)
            send_packet_name_to_player(packets,room_id)
            logger.info("get question correctly")

            return packets
            
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@test_routers.get("/api/test/question/round")
def get_each_round_questions(test_name:str, round: str, room_id: str,request: Request, packet_name: Optional[str] = None, difficulty: Optional[str] = None):
    # player_answer[f"{room_id}"] = []
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        result = process_test_data(authenticated_uid, test_name)

        try:
            logger.info("Attempting to create reference to Firebase")
            questions = get_questions_by_round(result,round,packet_name, difficulty)
            logger.info("get question correctly")
            
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    return questions

@test_routers.get("/api/test/question/prefetch")
def prefetch_question(test_name:str, round: str, room_id: str, request: Request, packet_name: Optional[str] = None, difficulty: Optional[str] = None, question_number: Optional[str] | None= None):
    """Prefetch next question without sending to players - for host preview"""
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        if not authenticated_uid:
            raise HTTPException(status_code=401, detail="Unauthorized: User ID not found")

        result = process_test_data(authenticated_uid, test_name)

        try:
            logger.info("Prefetching question for host preview")
            question_with_answer = get_specific_question(result, round, packet_name, difficulty, None, question_number)
            logger.info("Prefetch successful")

            # Return question with answer for host preview (don't send to Firebase)
            return {
                "question": question_with_answer,
                "answer": question_with_answer.get("answer", ""),
                "prefetch": True
            }

        except Exception as e:
            logger.error(f"Error prefetching question: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise HTTPException(status_code=500, detail=f"Error prefetching question: {str(e)}")

    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        logger.error(f"Error processing prefetch request: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@test_routers.get("/api/test/question")
def get_question_one_by_one(test_name:str, round: str, room_id: str,request: Request, packet_name: Optional[str] = None, difficulty: Optional[str] = None, question_number: Optional[str] | None= None, page: Optional[int] | None = None, limit: Optional[int] | None = None):

    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        logger.info(f"authenticated_uid: {authenticated_uid}")
        result = process_test_data(authenticated_uid, test_name)


        logger.info(f"result {result}")

        try:
            logger.info("Attempting to create reference to Firebase")

            question = get_specific_question(result,round,packet_name, difficulty, question_number=question_number, page=page, limit=limit)
            try:
                logger.info("Attempting to create reference to Firebase")
                logger.info(f"question_with_answer {question}")
                if isinstance(question, dict):
                    current_correct_answer = question.get("answer")
                    if current_correct_answer is not None:
                        current_correct_answer = str(current_correct_answer).split("~/")
                    set_current_correct_answer(room_id,current_correct_answer)
                    question_without_answer =  {key: value for key, value in question.items() if key != "answer"}


                elif isinstance(question, list):
                    question_without_answer = [
                        {key: value for key, value in q.items() if key != "answer"}
                        for q in question if isinstance(q, dict)
                    ]

                # Trường hợp không hợp lệ
                else:
                    raise TypeError(f"Expected dict or list of dicts, but got {type(question)}")

                logger.info(f"question_without_answer {question_without_answer}")
                
                send_question_to_player(room_id,question_without_answer)

                # Tiếp tục logic của bạn
            
            except Exception as e:
                logger.error(f"Error creating Firebase reference: {str(e)}")
                logger.error(f"Full traceback: {traceback.format_exc()}")
                raise  # Ném lại lỗi để xử lý ở tầng trên nếu cần

            return question

            

            # Tiếp tục logic của bạn
        
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise  # Ném lại lỗi để xử lý ở tầng trên nếu cần

        # try:
        #     logger.info("Attempting to create reference to Firebase")
        #     question_without_answer = {key: value for key , value in question.items() if key!="answer"}
        #     logger.info(f"question_with_answer {question_without_answer}")
            

        #     # Tiếp tục logic của bạn
        
        # except Exception as e:
        #     logger.error(f"Error creating Firebase reference: {str(e)}")
        #     logger.error(f"Full traceback: {traceback.format_exc()}")
        #     raise  # Ném lại lỗi để xử lý ở tầng trên nếu cần


        # try:
        #     logger.info("Attempting to create reference to Firebase")
        #     send_question_to_player(question_without_answer,room_id)
        #     logger.info("send successfully")
            

        #     # Tiếp tục logic của bạn
        
        # except Exception as e:
        #     logger.error(f"Error creating Firebase reference: {str(e)}")
        #     logger.error(f"Full traceback: {traceback.format_exc()}")
        #     raise  # Ném lại lỗi để xử lý ở tầng trên nếu cần
        
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    # return question

@test_routers.post("/api/test/answer")
def get_question_one_by_one(room_id: str,answer:Answer, request: Request):
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        if not authenticated_uid:
            logger.info(f"abc")
            raise HTTPException(status_code=401, detail="Unauthorized: User ID not found")
        try:
            logger.info("Attempting to send answer to player")
            send_answer_to_player(answer.answer,room_id)
            logger.info("send answer correctly")

        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise  
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/time")
def get_question_one_by_one(room_id: str,request: Request):
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        if not authenticated_uid:
            logger.info(f"abc")
            raise HTTPException(status_code=401, detail="Unauthorized: User ID not found")
        
        try:
            logger.info("Attempting to start time")
            start_time(room_id)
            logger.info("time started")

        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise  
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:

        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/submit")
def submit_answer(room_id: str, answer: Answer,request: Request):
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        player_answer = get_player_answer(room_id, authenticated_uid)

        logger.info(f"player {player_answer}")
        current_correct_answer = get_current_correct_answer(room_id)
        if not authenticated_uid:
            logger.info(f"abc")
            raise HTTPException(status_code=401, detail="Unauthorized: User ID not found")
    
        
        try:

            # Build base player object
            player_answer_object = {
                "uid": authenticated_uid,
                "player_name": answer.player_name,
                "avatar": answer.avatar,
                "answer": answer.answer,
                "stt": answer.stt,
                "row": answer.row,
                "time": float(answer.time),
                "isObstacle": answer.is_obstacle,
                "round_scores": {
                    "1": 0,
                    "2": 0,
                    "3": 0,
                    "4": 0
                },
                "was_deducted_this_round": False ,
                "is_correct": False  # Always start with False, will update if correct
            }

            # Check if this player has submitted before
            # existing_player = None
            # for player in player_answer:
            #     if player["uid"] == authenticated_uid:
            #         existing_player = player
            #         break  
            # logger.info(f"existing_player {existing_player}")

            # if existing_player:
            #     # Update existing player object
            #     existing_player.update({
            #         "answer": answer.answer,
            #         "stt": answer.stt,
            #         "row": answer.row,
            #         "time": float(answer.time),
            #         "isObstacle": answer.is_obstacle,
            #         "is_correct": False  # reset is_correct; will update below if needed
            #     })
            # else:
            #     # New player submission
            #     player_answer_object["score"] = 0
            #     player_answer.append(player_answer_object)


            submitted = normalize_string(answer.answer)



            logger.info(f"player answer {player_answer}")

            logger.info(f"submit {submitted}")
            logger.info(f"correcr {current_correct_answer}")

            player_answer["answer"] = answer.answer
            player_answer["time"] = float(answer.time)

            if any(submitted == normalize_string(correct_answer) for correct_answer in current_correct_answer):
                logger.info(f"submit {submitted}")
                player_answer["is_correct"] = True


            logger.info(f"player_answer {player_answer}")

            set_single_player_answer(room_id, authenticated_uid, player_answer)


        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise  
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:

        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/score/rules")
def apply_score_rule(room_id: str, rules: ScoreRule, request: Request):
    try:
        try:
            score_rules[room_id] = rules.dict()
            send_score_rule(room_id, rules.dict())

        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        

    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:

        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/score")
def send_score_to_player(
        room_id: str, 
        mode: str,
        request: Request,
        scores: Optional[List[Score]] = None, 
        round: Optional[str] = None, 
        stt: Optional[str] = None,
        is_obstacle_correct: Optional[str] = None,
        obstacle_point: Optional[int] = None,
        is_correct: Optional[str] = None,
        round_4_mode: Optional[str] = None,
        difficulty: Optional[str] = None,
        is_take_turn_correct: Optional[str] = None,
        stt_take_turn: Optional[str] = None,
        stt_taken: Optional[str] = None 
    ):
    try:
        try:

            user = request.state.user
            authenticated_uid = user["uid"]
            score_list = []
            score_rules = get_score_rules(room_id)
            player_answer = get_all_player_answer(room_id)
            logger.info(f"player answer in scoring {player_answer}")

            # for player in player_answer:
            #     if "round_scores" not in player:
            #         player["round_scores"] = {} 
            # }

            if is_obstacle_correct:
                for player in player_answer:
                    if player["stt"] == stt:
                        player["score"] += obstacle_point
                        player["round_scores"][int(round)] += obstacle_point
                        set_player_answer(room_id, player["uid"], player)
                        logger.info(f"[Round4-Main] {player['uid']} +{obstacle_point}")
                
                    score_list.append({
                        "playerName": player["userName"],
                        "avatar": player["avatar"],
                        "score": player["score"],
                        "isCorrect": str(player["is_correct"]).lower(),
                        "isModified": "true",
                        "stt": player["stt"]
                    })

                
            if mode == "manual" and not is_obstacle_correct:
                for score in scores:
                    score_list.append({
                        "playerName": score.playerName,
                        "avatar": score.avatar,
                        "score": score.score,
                        "isCorrect": score.isCorrect,
                        "isModified": score.isModified,
                        "stt": score.stt
                    })

                    # round_scores.append({
                    #     "stt": score.stt,
                    #     "playerName": score.playerName,
                    #     "roundScore": score.score
                    # })


            if mode == "adaptive" and not is_obstacle_correct:
                correct_players = [p for p in player_answer if p.get("is_correct") == True]
                correct_count = len(correct_players)

                if correct_count == 4:
                    points = 5
                elif correct_count == 3:
                    points = 10
                elif correct_count == 2:
                    points = 15
                elif correct_count == 1:
                    points = 20
                else:
                    points = 0

                logger.info(f"[Round {round} - correct_count_bonus] {correct_count} correct players, awarding {points} points each")

                for player in correct_players:
                    logger.info(f'player["round_scores"] {player["round_scores"]}')
                    player["score"] += points
                    player["round_scores"][int(round)] = player["round_scores"][int(round)] + points
                    logger.info(f"[Round {round} - correct_count_bonus] {player['uid']} +{points}")
                    set_player_answer(room_id, player["uid"], player)

                try:            
                    for player in player_answer:
                        # Check if this player is correct this round
                        # if player["is_correct"] and mode!="adaptive":
                        #     index = next((i for i, p in enumerate(correct_players) if p["uid"] == player["uid"]), None)
                        #     bonus = score_rules[room_id][f"round{round}"][index] if index is not None else 0
                        #     player["score"] += bonus

                        score_list.append({
                            "playerName": player["userName"],
                            "avatar": player["avatar"],
                            "score": player["score"],
                            "isCorrect": str(player["is_correct"]).lower(),
                            "isModified": "true",
                            "stt": player["stt"]
                        })

                except Exception as e:
                    logger.error(f"Error creating Firebase reference: {str(e)}")
                    logger.error(f"Full traceback: {traceback.format_exc()}")
                    raise

            if mode == "auto" and not is_obstacle_correct:
                try:
                    logger.info("Attempting to submit answer")
                    correct_players = sorted(
                        [p for p in player_answer if p.get("is_correct") == True],
                        key=lambda x: x["time"]
                    )

                    logger.info(f"correct_players {correct_players}")
                    logger.info(f"rules {score_rules}")

                except Exception as e:
                    logger.error(f"Error creating Firebase reference: {str(e)}")
                    logger.error(f"Full traceback: {traceback.format_exc()}")
                    raise

                try:

                    if round == "3" and stt and is_correct and is_correct.lower() == "true":

                        matched_player = next(
                            (p for p in player_answer if p["stt"] == stt),
                            None
                        )
                        if matched_player:
                            matched_player["score"] += score_rules[f"round{round}"]
                            matched_player["round_scores"][int(round)] += score_rules[f"round{round}"]
                            matched_player["is_correct"] = True
                            set_player_answer(room_id, matched_player["uid"], matched_player)
                            logger.info(f"Updated player {matched_player['uid']} score for round 3")

                except Exception as e:
                    logger.error(f"Error creating Firebase reference: {str(e)}")
                    logger.error(f"Full traceback: {traceback.format_exc()}")
                    raise

                try:

                    if round == "4" and round_4_mode and difficulty:
                        logger.info(f"difficulty {difficulty}")
                        diff_index = {"Dễ": 0, "Trung bình": 1, "Khó": 2}.get(difficulty)
                        if diff_index is None:
                            raise HTTPException(status_code=400, detail="Invalid difficulty")

                        points = score_rules["round4"][diff_index]

                        if round_4_mode == "main":
                            for player in player_answer:
                                if player["stt"] == stt:
                                    player["score"] += points
                                    player["round_scores"][int(round)] += points
                                    set_player_answer(room_id, player["uid"], player)
                                    logger.info(f"[Round4-Main] {player['uid']} +{points}")

                        elif round_4_mode == "nshv":
                            for player in player_answer:
                                if player["stt"] == stt:
                                    if is_correct == "true":
                                        added = int(points * 1.5)
                                        player["score"] += added
                                        player["round_scores"][int(round)] += added
                                        set_player_answer(room_id, player["uid"], player)
                                        logger.info(f"[Round4-NSHV] {player['uid']} +{added}")
                                    if is_correct == "false":
                                        deducted = points 
                                        player["score"] -= deducted
                                        player["round_scores"][int(round)] -= deducted
                                        player["was_deducted_this_round"] = True
                                        set_player_answer(room_id, player["uid"], player)
                                        logger.info(f"[Round4-NSHV] {player['uid']} -{deducted}")

                        elif round_4_mode == "take_turn":
                            if is_take_turn_correct == "false":
                                for player in player_answer:
                                    if player["stt"] == stt_take_turn:
                                        deducted = points // 2
                                        player["score"] = max(0, player["score"] - deducted)
                                        player["round_scores"][int(round)] -= deducted
                                        set_player_answer(room_id, player["uid"], player)
                                        logger.info(f"[Round4-TakeTurn-False] {player['uid']} -{deducted}")
                            elif is_take_turn_correct == "true":
                                taker = next((p for p in player_answer if p["stt"] == stt_take_turn), None)
                                taken = next((p for p in player_answer if p["stt"] == stt_taken), None)

                                if taker and taken:
                                    if not taken.get("was_deducted_this_round"):                
                                        deducted = points 
                                        taken["score"] = max(0, taken["score"] - deducted)
                                        taken["round_scores"][int(round)] -= deducted
                                        set_player_answer(room_id, taken["uid"], taken)
                                        taker["score"] += points
                                        taker["round_scores"][int(round)] +=points
                                        set_player_answer(room_id, taker["uid"], taker)
                                        logger.info(f"[Round4-TakeTurn-True] {taker['uid']} +{points}, {taken['uid']} -{deducted}")

                        else:
                            raise HTTPException(status_code=400, detail="Invalid round_4_mode")

                except Exception as e:
                    logger.error(f"Error creating Firebase reference: {str(e)}")
                    logger.error(f"Full traceback: {traceback.format_exc()}")
                    raise

                try:            
                    for player in player_answer:
                        # Check if this player is correct this round
                        if player["is_correct"] and mode!="adaptive":
                            index = next((i for i, p in enumerate(correct_players) if p["uid"] == player["uid"]), None)
                            bonus = score_rules[f"round{round}"][index] if index is not None else 0
                            player["score"] += bonus
                            player["round_scores"][int(round)] += bonus
                            set_player_answer(room_id, player["uid"], player)

                        score_list.append({
                            "playerName": player["userName"],
                            "avatar": player["avatar"],
                            "score": player["score"],
                            "isCorrect": str(player["is_correct"]).lower(),
                            "isModified": "true",
                            "stt": player["stt"]
                        })

                except Exception as e:
                    logger.error(f"Error creating Firebase reference: {str(e)}")
                    logger.error(f"Full traceback: {traceback.format_exc()}")
                    raise
                
                
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise 
        
        if not authenticated_uid:
            logger.info(f"abc")
            raise HTTPException(status_code=401, detail="Unauthorized: User ID not found")
        
        try:
            logger.info("Attempting to submit answer")
            send_score(room_id,mode, score_list)
            logger.info("time started")

        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise

        try:
            logger.info("Attempting to submit answer")
            for player in player_answer:
                player["is_correct"] = False
                player["was_deducted_this_round"] = False
                set_player_answer(room_id, player["uid"], player)
                logger.info(f"player {player}")

            
            firebase_data = {
                player["stt"]: {
                    "playerName": player["userName"],
                    "avatar": player["avatar"],
                    "roundScore": player["round_scores"][int(round)]
                }
                for player in player_answer
            }
            update_score_each_round(room_id,firebase_data,round)
        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise

         

    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:

        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    
@test_routers.post("/api/test/broadcast")
def broadcast_answer(room_id: str,request: Request):
    try:
        user = request.state.user
        authenticated_uid = user["uid"]
        player_answer = get_all_player_answer(room_id)
        if not authenticated_uid:
            logger.info(f"abc")
            raise HTTPException(status_code=401, detail="Unauthorized: User ID not found")
        
        try:
            logger.info("Attempting to submit answer")
            broadcast_player_answer(room_id,player_answer)
            logger.info("time started")

        except Exception as e:
            logger.error(f"Error creating Firebase reference: {str(e)}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise  

        # player_answer = []

    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:

        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    return player_answer



@test_routers.get("/api/test/{test_name}")
def read_test(test_name: str, request: Request):
    user = request.state.user
    authenticated_uid = user["uid"]
    test = get_test_by_name(authenticated_uid, test_name)[0]["questions"]
    logger.info(f"test: {test}")
    grouped_round_3 = defaultdict(list)
    for question in [question for question in test if question["round"] == 3]:
        grouped_round_3[question["packetName"]].append(question)
    
    for packet_name in grouped_round_3:
        grouped_round_3[packet_name].sort(key=lambda x: x["stt"])

    grouped_round_4 = defaultdict(list)
    for question in [question for question in test if question["round"] == 4]:
        grouped_round_4[question["difficulty"]].append(question)
    
    for packet_name in grouped_round_4:
        grouped_round_4[packet_name].sort(key=lambda x: x["stt"])

    round_1 = sorted([question for question in test if question["round"] == 1]  , key=lambda x: x["stt"]) #get question and sort by stt
    round_2 = sorted([question for question in test if question["round"] == 2]  , key=lambda x: x["stt"])
    round_3 = dict(grouped_round_3)
    round_4 = dict(grouped_round_4)
    turn = sorted([question for question in test if question["round"] == "turn"]  , key=lambda x: x["stt"])
    
    result = {
        "round_1": round_1,
        "round_2": round_2,
        "round_3": round_3,
        "round_4": round_4,
        "turn": turn
    }

    if "error" in test:
        raise HTTPException(status_code=500, detail=test["error"])
    if len(test) == 0:
        raise HTTPException(status_code=404, detail="Test not found")
    return result


@test_routers.post("/api/test/upload")
async def process_file(test_name: str , request: Request, file: UploadFile = File(...)):
    try:
        # Kiểm tra xem test_name đã tồn tại chưa
        user = request.state.user
        authenticated_uid = user["uid"]
        print(f"user: {user}")
        logger.info(f"authenticated_uid: {authenticated_uid}")
        existing_test = db.collection("tests").where("testName", "==", test_name).limit(1).get()
        if existing_test:
            raise HTTPException(status_code=400, detail=f"Bộ đề với tên '{test_name}' đã tồn tại.")

        # Tạo một document trong "tests" cho toàn bộ file upload
        test_ref = db.collection("tests").document()  
        test_id = test_ref.id
        test_data = {
            "testId": test_id,
            "testName": test_name,
            "createdAt": firestore.SERVER_TIMESTAMP,
            "owner": authenticated_uid,
            "totalQuestions": 0,  
            "status": "active"
        }
        test_ref.set(test_data)
        # Đọc nội dung file vào bộ nhớ
        contents = await file.read()

        # Tải file Excel từ nội dung trong bộ nhớ
        workbook = load_workbook(BytesIO(contents))


        # Kết quả tổng hợp từ tất cả các sheet
        result = {
            "filename": file.filename,
            "sheets": {}
        }

        # Duyệt qua từng sheet trong workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logger.info(f"sheet: {sheet}")
            data = []

            # Đọc từng dòng trong sheet
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # Xử lý dữ liệu theo tên sheet bằng match-case
            match sheet_name.lower():  # Chuyển thành chữ thường để tránh lỗi
                case "round1":
                    # Xử lý cho Round 1 (ví dụ: chỉ có question và answer)
                    processed_data = [
                        {
                            "stt": int(row[0]) if row[0] else None,  # Cột STT
                            "question": row[1],                      # Cột Question
                            "answer": row[2], 
                            "type": row[3],                          # Cột Answer
                            "imgUrl": row[4]                         # Cột ImgUrl
                        }
                        for row in data[1:]  # Bỏ header
                        if row[1] and row[2]  # Kiểm tra question và answer không null
                    ]

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["round1"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": 1
                    }

                    print(processed_data)

                    # Gọi hàm upload ngay sau khi xử lý dữ liệu
                    if processed_data:  # Chỉ upload nếu có dữ liệu
                        upload_result = upload_test_to_firestore(
                            rounds=1, 
                            questions=processed_data,  # Dữ liệu câu hỏi đã xử lý
                            test_id=test_id
                        )
                        result["sheets"]["round1"]["upload_result"] = upload_result
                    else:
                        result["sheets"]["round1"]["upload_result"] = {
                            "message": "Không có dữ liệu hợp lệ để upload cho sheet round1."
                        }

                case "round2":

                    # Xử lý cho Round 2 (ví dụ: thêm images)
                    processed_data = [
                        {
                            "stt": int(row[0]) if row[0] else None,  # Cột STT,
                            "question": row[1],                      # Cột Question
                            "answer": row[2], 
                        }
                        for row in data[1:] if row[1] and row[2] #check null
                    ]

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["round2"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": 2
                    }

                    if processed_data:  # Chỉ upload nếu có dữ liệu
                        upload_result = upload_test_to_firestore(
                            rounds=2, 
                            questions=processed_data,  # Dữ liệu câu hỏi đã xử lý
                            test_id=test_id
                        )
                        result["sheets"]["round2"]["upload_result"] = upload_result
                    else:
                        result["sheets"]["round2"]["upload_result"] = {
                            "message": "Không có dữ liệu hợp lệ để upload cho sheet round2."
                        }

                case "round3":
                    # Xử lý cho Round 1 (ví dụ: chỉ có question và answer)
                    packet_name = ""
                    # Xử lý dữ liệu từ sheet
                    processed_data = []
                    for row in data[1:]:  # Bỏ header
                        if row[0] == "Tên gói":  
                            packet_name = row[1]  
                            continue  
                        
                        if row[1] and row[2]:
                            processed_data.append({
                                "stt": int(row[0]) if row[0] else None,  # Cột STT
                                "question": row[1],                      # Cột Question
                                "answer": row[2],                        # Cột Answer
                                "packetName": packet_name               # Thêm packet_name vào object
                            })

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["round3"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": 3
                    }

                    print(processed_data)

                    # Gọi hàm upload ngay sau khi xử lý dữ liệu
                    if processed_data:  # Chỉ upload nếu có dữ liệu
                        upload_result = upload_test_to_firestore(
                            rounds=3, 
                            questions=processed_data,  # Dữ liệu câu hỏi đã xử lý
                            test_id=test_id
                        )
                        result["sheets"]["round3"]["upload_result"] = upload_result
                    else:
                        result["sheets"]["round3"]["upload_result"] = {
                            "message": "Không có dữ liệu hợp lệ để upload cho sheet round3."
                        }

                case "round4":
                    difficulty = ""
                    # Xử lý dữ liệu từ sheet
                    processed_data = []
                    for row in data[1:]:  # Bỏ header
                        logger.info(f"row: {row}")
                        if row[0] == "Mức độ":  
                            difficulty = row[1] 
                            continue  
                        
                        if row[1] and row[2]:
                            processed_data.append({
                                "stt": int(row[0]) if row[0] else None,  # Cột STT
                                "question": row[1],                      # Cột Question
                                "answer": row[2], 
                                "url": row[3],                       # Cột Answer
                                "difficulty": difficulty               
                            })

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["round4"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": 4
                    }

                    print(processed_data)

                    # Gọi hàm upload ngay sau khi xử lý dữ liệu
                    
                    if processed_data:  # Chỉ upload nếu có dữ liệu
                        upload_result = upload_test_to_firestore(
                            rounds=4, 
                            questions=processed_data,  # Dữ liệu câu hỏi đã xử lý
                            test_id=test_id
                        )
                        result["sheets"]["round4"]["upload_result"] = upload_result
                    else:
                        result["sheets"]["round4"]["upload_result"] = {
                            "message": "Không có dữ liệu hợp lệ để upload cho sheet round4."
                        }
                case "turn":
                    # Xử lý cho phân lượt (ví dụ: chỉ có question và answer)
                    processed_data = [
                        {
                            "stt": int(row[0]) if row[0] else None,  # Cột STT
                            "question": row[1],                      # Cột Question
                            "answer": row[2], 
                        }
                        for row in data[1:]  # Bỏ header
                        if row[1] and row[2]  # Kiểm tra question và answer không null
                    ]

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["turn"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": "turn"
                    }

                    print(processed_data)

                    # Gọi hàm upload ngay sau khi xử lý dữ liệu
                    if processed_data:  # Chỉ upload nếu có dữ liệu
                        upload_result = upload_test_to_firestore(
                            rounds="turn", 
                            questions=processed_data,  # Dữ liệu câu hỏi đã xử lý
                            test_id=test_id
                        )
                        result["sheets"]["turn"]["upload_result"] = upload_result
                    else:
                        result["sheets"]["turn"]["upload_result"] = {
                            "message": "Không có dữ liệu hợp lệ để upload cho sheet turn."
                        }


                case _:
                    # Xử lý mặc định nếu sheet không khớp
                    logger.warning(f"Sheet không xác định: {sheet_name}")
                    result["sheets"][sheet_name] = {
                        "sheet_name": sheet_name,
                        "content": data,
                        "round": None
                    }

            logger.info(f"Đã xử lý sheet: {sheet_name} với {len(data)} dòng")

        return result

    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

    finally:
        await file.close()


@test_routers.post("/api/test/question/add")
async def add_new_question_to_test(question: UpdateQuestionRequest):
    try:

        await upload_single_question_to_firestore(question)
        
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
