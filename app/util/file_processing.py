from fastapi import HTTPException, logger
from fastapi import FastAPI, Depends, UploadFile
from openpyxl import load_workbook
from io import BytesIO
from ..repositories.firestore.test_repository import TestRepository


async def process_excel_file(test_id: str, file: UploadFile, test_repository: TestRepository):

    contents = await file.read()
    # Tải file Excel từ nội dung trong bộ nhớ
    workbook = load_workbook(BytesIO(contents))

    # Kết quả tổng hợp từ tất cả các sheet
    result = {
        "filename": file.filename,
        "sheets": {}
    }
    try:
        # Duyệt qua từng sheet trong workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logger.info(f"sheet: {sheet}")
            data = []

            # Đọc từng dòng trong sheet
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # Xử lý dữ liệu theo tên sheet bằng match-case
            match sheet_name.lower():  # Chuyển thành chữ thường để tránh lỗi
                case "round1":
                    # Xử lý cho Round 1 (ví dụ: chỉ có question và answer)
                    processed_data = [
                        {
                            "stt": int(row[0]) if row[0] else None,  # Cột STT
                            "question": row[1],                      # Cột Question
                            "answer": row[2], 
                            "type": row[3],                          # Cột Answer
                            "imgUrl": row[4]                         # Cột ImgUrl
                        }
                        for row in data[1:]  # Bỏ header
                        if row[1] and row[2]  # Kiểm tra question và answer không null
                    ]

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["round1"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": 1
                    }

                    print(processed_data)

                    process_sheet(processed_data, test_id, "1", test_repository)

                case "round2":

                    # Xử lý cho Round 2 (ví dụ: thêm images)
                    processed_data = [
                        {
                            "stt": int(row[0]) if row[0] else None,  # Cột STT,
                            "question": row[1],                      # Cột Question
                            "answer": row[2], 
                        }
                        for row in data[1:] if row[1] and row[2] #check null
                    ]

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["round2"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": 2
                    }

                    process_sheet(processed_data, test_id, "2", test_repository)

                case "round3":
                    # Xử lý cho Round 3 (ví dụ: chỉ có question và answer)
                    packet_name = ""
                    # Xử lý dữ liệu từ sheet
                    processed_data = []
                    for row in data[1:]:  # Bỏ header
                        if row[0] == "Tên gói":  
                            packet_name = row[1]  
                            continue  
                        
                        if row[1] and row[2]:
                            processed_data.append({
                                "stt": int(row[0]) if row[0] else None,  # Cột STT
                                "question": row[1],                      # Cột Question
                                "answer": row[2],                        # Cột Answer
                                "packetName": packet_name               # Thêm packet_name vào object
                            })

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["round3"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": 3
                    }

                    print(processed_data)

                    process_sheet(processed_data, test_id, "3", test_repository)

                case "round4":
                    # Đọc 60 câu hỏi liên tục và đánh ID theo thứ tự
                    processed_data = []
                    question_count = 0

                    for row in data[1:]:  # Bỏ header
                        logger.info(f"row: {row}")

                        if row[1] and row[2]:  # Có question và answer
                            question_count += 1
                            processed_data.append({
                                "stt": question_count,  # ID liên tục từ 1-60
                                "question": row[1],     # Cột Question
                                "answer": row[2],       # Cột Answer
                                "url": row[3] if len(row) > 3 else None,  # Cột URL (nếu có)
                                "difficulty": None      # Sẽ được xác định động dựa trên cấu hình phòng
                            })

                            if question_count >= 60:
                                break

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["round4"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": 4
                    }

                    print(processed_data)
                    process_sheet(processed_data, test_id, "4", test_repository)
                case "turn":
                    # Xử lý cho phân lượt (ví dụ: chỉ có question và answer)
                    processed_data = [
                        {
                            "stt": int(row[0]) if row[0] else None,  # Cột STT
                            "question": row[1],                      # Cột Question
                            "answer": row[2], 
                        }
                        for row in data[1:]  # Bỏ header
                        if row[1] and row[2]  # Kiểm tra question và answer không null
                    ]

                    logger.info(f"processed_data: {processed_data}")
                    result["sheets"]["turn"] = {
                        "sheet_name": sheet_name,
                        "content": processed_data,
                        "round": "turn"
                    }

                    print(processed_data)
                    process_sheet(processed_data, test_id, "turn", test_repository)

                case _:
                    logger.warning(f"Sheet không xác định: {sheet_name}")
                    result["sheets"][sheet_name] = {
                        "sheet_name": sheet_name,
                        "content": data,
                        "round": None
                    }
            logger.info(f"Đã xử lý sheet: {sheet_name} với {len(data)} dòng")

        return result
    
    except HTTPException as http_exc:
        raise http_exc  # Re-raise HTTPException to let FastAPI set the status code
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    finally:
        await file.close()


def process_sheet(processed_data: list, test_id: str, round: str, test_repository: TestRepository):
    if processed_data:  # Chỉ upload nếu có dữ liệu
        upload_result = test_repository.set_test_by_batch(
            questions=processed_data,  # Dữ liệu câu hỏi đã xử lý
            test_id=test_id,
            round=round
        )
        return upload_result
    else:
        return {
            "message": f"Không có dữ liệu hợp lệ để upload cho sheet round{round}."
        }
